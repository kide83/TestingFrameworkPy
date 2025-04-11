from selenium.webdriver.chrome.options import Options
from selenium import webdriver
from selenium.webdriver.common.by import By
from utils.logger import get_logger
from selenium.webdriver.support.ui import Select
import os
import time
import random
import pandas as pd
from openpyxl import load_workbook

logger = get_logger()

class BrowserFunctions:
    def __init__(self, browser_name="Chrome"):
        options = Options()
        #options.add_argument("--headless")  # Run in headless mode (no UI)
        options.add_argument("--no-sandbox")  # Bypass OS security model
        options.add_argument("--disable-dev-shm-usage")  # Prevent /dev/shm issues
        options.add_argument("--user-data-dir=/tmp/chrome-user-data")  # Set a unique user data directory

        """Initialize the browser based on the given name."""
        if browser_name.lower() == "chrome":
            logger.info("Starting Chrome WebDriver")
            self.driver = webdriver.Chrome(options=options)
            self.driver.maximize_window()
            self.driver.implicitly_wait(10)
        elif browser_name.lower() == "edge":
            logger.info("Starting Edge WebDriver")
            self.driver = webdriver.Edge(options=options)
            self.driver.maximize_window()
            self.driver.implicitly_wait(10)
        else:
            raise ValueError(f"Unsupported browser: {browser_name}")

    def open_url(self, url):
        """Open a URL in the selected browser."""
        self.driver.get(url)

    def close_browser(self):
        """Close the browser instance."""
        logger.info("Closing WebDriver")
        self.driver.quit()

    def get_page_title(self):
        """Get the title of the current page."""
        return self.driver.title

    def check_dropdown_list_item(self, locator, locator_type, item_value):
        if locator_type.lower() == "id":
            dropdown_element = self.driver.find_element(By.ID, locator)
        elif locator_type.lower() == "xpath":
            dropdown_element = self.driver.find_element(By.XPATH, locator)
        else:
            raise ValueError("Invalid locator type! Use 'id' or 'xpath'.")
        dropdown = Select(dropdown_element)
        options = [option.text for option in dropdown.options]
        expected_option = item_value
        assert expected_option in options, f"'{expected_option}' not found in dropdown options"

    def click_dropdown_list_item(self, locator, locator_type, item_value):
        if locator_type.lower() == "id":
            dropdown_element = self.driver.find_element(By.ID, locator)
        elif locator_type.lower() == "xpath":
            dropdown_element = self.driver.find_element(By.XPATH, locator)
        else:
            raise ValueError("Invalid locator type! Use 'id' or 'xpath'.")
        dropdown = Select(dropdown_element)
        dropdown.select_by_visible_text(item_value)
        selected_option = dropdown.first_selected_option.text
        assert selected_option == item_value, f"Expected '{item_value}', but got '{selected_option}'"

    def click_random_dropdown_list_item(self, locator, locator_type):
        if locator_type.lower() == "id":
            dropdown_element = self.driver.find_element(By.ID, locator)
        elif locator_type.lower() == "xpath":
            dropdown_element = self.driver.find_element(By.XPATH, locator)
        else:
            raise ValueError("Invalid locator type! Use 'id' or 'xpath'.")
        dropdown = Select(dropdown_element)
        options = dropdown.options

        if len(options) > 1:
            random_choice = random.choice(options[1:])  # Skip the first (usually "Select" option)
            dropdown.select_by_visible_text(random_choice.text)
        else:
            raise ValueError("Dropdown does not have enough options to select from.")

    def click_element(self, locator, locator_type):
        if locator_type.lower() == "id":
            element = self.driver.find_element(By.ID, locator)
        elif locator_type.lower() == "xpath":
            element = self.driver.find_element(By.XPATH, locator)
        else:
            raise ValueError("Invalid locator type! Use 'id' or 'xpath'.")
        element.click()

    def get_element_text(self, locator, locator_type):
        if locator_type.lower() == "id":
            element = self.driver.find_element(By.ID, locator)
        elif locator_type.lower() == "xpath":
            element = self.driver.find_element(By.XPATH, locator)
        else:
            raise ValueError("Invalid locator type! Use 'id' or 'xpath'.")
        return element.text

    def take_screenshot(self, filename):
        """Captures a screenshot of the current page."""
        timestamp = time.strftime("%Y-%m-%d_%H-%M-%S")
        screenshot_name = f"screenshots/{filename}_{timestamp}.png"

        os.makedirs("screenshots", exist_ok=True)

        self.driver.save_screenshot(screenshot_name)
        logger.info(f"Screenshot saved as {filename}")

    def enter_text(self, locator, locator_type, text):
        if locator_type.lower() == "id":
            element = self.driver.find_element(By.ID, locator)
        elif locator_type.lower() == "xpath":
            element = self.driver.find_element(By.XPATH, locator)
        else:
            raise ValueError("Invalid locator type! Use 'id' or 'xpath'.")
        element.clear()
        element.send_keys(text)

    def click_button(self, locator, locator_type):
        if locator_type.lower() == "id":
            button = self.driver.find_element(By.ID, locator)
        elif locator_type.lower() == "xpath":
            button = self.driver.find_element(By.XPATH, locator)
        else:
            raise ValueError("Invalid locator type! Use 'id' or 'xpath'.")
        button.click()

    def select_checkbox(self, locator, locator_type):
        if locator_type.lower() == "id":
            checkbox = self.driver.find_element(By.ID, locator)
        elif locator_type.lower() == "xpath":
            checkbox = self.driver.find_element(By.XPATH, locator)
        else:
            raise ValueError("Invalid locator type! Use 'id' or 'xpath'.")
        checkbox.click()

    def select_radio_button(self, group_name, value):
        radio_buttons = self.driver.find_elements(By.NAME, group_name)
        for radio in radio_buttons:
            if radio.get_attribute("value") == value:
                radio.click()
                return
        raise ValueError(f"Radio button with value '{value}' not found.")
### Treba promeniti kod po potrebi, mozda da koristi ID ili xpath i isprobati na realnom promeru...

    def get_table_cell_value(self, table_id, row_index, col_index):
        cell_xpath = f"//*[@id='{table_id}']/tbody/tr[{row_index}]/td[{col_index}]"
        cell = self.driver.find_element(By.XPATH, cell_xpath)
        return cell.text

    def click_button_in_table_cell(self, table_id, row_index, button_class):
        """Clicks a button inside a table at a specific row index."""
        table = self.driver.find_element(By.ID, table_id)
        rows = table.find_elements(By.TAG_NAME, "tr")

        if 0 < row_index < len(rows):
            button = rows[row_index].find_element(By.CLASS_NAME, button_class)
            button.click()
        else:
            raise IndexError("Row index out of range")

    # Required for function is_value_found_in_table()
    def get_table_data(self, table_id):
        """Extracts all rows from a table and returns them as a list of lists."""
        table = self.driver.find_element(By.ID, table_id)
        rows = table.find_elements(By.TAG_NAME, "tr")  # Get all table rows

        table_data = []
        for row in rows:
            cells = row.find_elements(By.TAG_NAME, "td")  # Get all cells in row
            row_data = [cell.text.strip() for cell in cells]  # Extract text
            if row_data:  # Avoid empty rows (like headers)
                table_data.append(row_data)
        return table_data

    def is_value_found_in_table(self, table_id, value):
        """Checks if a value exists in a table."""
        table_data = self.get_table_data(table_id)
        for row in table_data:
            if value in row:
                return True
        return False

    def get_column_number(self, table_id, header_name):
        """Finds the column number (index) of a given header in a table."""
        table = self.driver.find_element(By.ID, table_id)
        header_row = table.find_element(By.TAG_NAME, "tr")  # Get the first row (headers)
        headers = header_row.find_elements(By.TAG_NAME, "th")  # Get all header columns

        for index, header in enumerate(headers):
            if header.text.strip() == header_name:
                return index + 1  # Return 1-based index (like human counting)

        return -1  # Return -1 if header is not found

    def get_rows_count(self, table_id):
        """Returns the number of rows in the table (excluding headers)."""
        table = self.driver.find_element(By.ID, table_id)
        rows = table.find_elements(By.TAG_NAME, "tr")  # Get all rows
        return len(rows) - 1  # Subtract 1 to exclude header row



class OtherFunctions:
    def __init__(self):
        print()

    @staticmethod
    def generate_random_number(min_value, max_value):
        return random.randint(min_value, max_value)

    def generate_random_vorschrift_number(self):
        random_number = self.generate_random_number(1001, 9999)
        return f"(QS) {random_number}/{random_number + 3}"

    def generate_random_arbeitskreiss_number(self):
        random_number = self.generate_random_number(1001, 9999)
        return f"AK_QSIT_{random_number}"

    def get_value_from_excel(self, file_path, sheet_name, column_name, row_index):
        """Reads an Excel file and gets a value from a specific column and row in a given sheet."""
        # row_index = 2 je prvi red, red 1 je headers
        self.file_path = file_path
        self.workbook = load_workbook(file_path)
        sheet = self.workbook[sheet_name]
        headers = [cell.value for cell in sheet[1]]
        if column_name in headers:
            col_index = headers.index(column_name) + 1
            return sheet.cell(row=row_index, column=col_index).value
        else:
            raise Exception(f"Column '{column_name}' not found in sheet '{sheet_name}'")

    def write_to_excel(self, file_path, sheet_name, column_name, row_index, value):
        self.file_path = file_path
        self.workbook = load_workbook(file_path)
        sheet = self.workbook[sheet_name]
        headers = [cell.value for cell in sheet[1]]
        if column_name in headers:
            col_index = headers.index(column_name) + 1
            sheet.cell(row=row_index, column=col_index).value = value
            self.workbook.save(self.file_path)
        else:
            raise Exception(f"Column '{column_name}' not found in sheet '{sheet_name}'")











